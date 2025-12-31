from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Depends
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from fastapi.responses import StreamingResponse
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import pandas as pd
import io
import hashlib
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBasic()

# ================== MODELS ==================

class UserCreate(BaseModel):
    password: str

class LoginRequest(BaseModel):
    password: str

class TokenResponse(BaseModel):
    token: str
    message: str

# Categories (hierarchical)
class CategoryCreate(BaseModel):
    name: str
    parent_id: Optional[str] = None  # For sub-categories
    type: str  # income, expense
    icon: Optional[str] = None
    color: Optional[str] = None

class Category(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    parent_id: Optional[str] = None
    type: str  # income, expense
    icon: Optional[str] = None
    color: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Accounts (Bank, Cash, Credit Card, etc.)
class AccountCreate(BaseModel):
    name: str
    account_type: str  # bank, cash, credit_card, investment, loan_receivable, loan_payable
    opening_balance: float = 0.0
    description: Optional[str] = ""
    person_name: Optional[str] = None  # For loan accounts

class Account(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    account_type: str
    opening_balance: float = 0.0
    current_balance: float = 0.0
    description: str = ""
    person_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Transactions with proper double-entry
class TransactionCreate(BaseModel):
    date: str
    description: str
    amount: float
    account_id: str  # Source account (bank/cash)
    category_id: Optional[str] = None  # For income/expense categorization
    payee_id: Optional[str] = None  # For transfers to another account/person
    transaction_type: str  # expense, income, transfer
    reference: Optional[str] = ""
    notes: Optional[str] = ""

class Transaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    description: str
    amount: float
    account_id: Optional[str] = None  # Source account (new field)
    ledger_id: Optional[str] = None  # Legacy field for backward compatibility
    category_id: Optional[str] = None
    payee_id: Optional[str] = None  # For transfers
    transaction_type: str  # expense, income, transfer
    reference: str = ""
    notes: str = ""
    source: str = "manual"  # manual, bank_import
    tag: Optional[str] = None  # Legacy field
    transfer_pair_id: Optional[str] = None  # Links paired transfer transactions
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    
    @property
    def effective_account_id(self):
        """Return account_id or fall back to ledger_id for legacy data"""
        return self.account_id or self.ledger_id

class TransactionUpdate(BaseModel):
    date: Optional[str] = None
    description: Optional[str] = None
    amount: Optional[float] = None
    account_id: Optional[str] = None
    category_id: Optional[str] = None
    payee_id: Optional[str] = None
    transaction_type: Optional[str] = None
    notes: Optional[str] = None

# For bank statement upload - tagging
class TagTransactionRequest(BaseModel):
    category_id: Optional[str] = None
    payee_id: Optional[str] = None

class BulkTagRequest(BaseModel):
    transaction_ids: List[str]
    category_id: Optional[str] = None
    payee_id: Optional[str] = None

# Loans
class LoanCreate(BaseModel):
    person_name: str
    loan_type: str  # given, taken
    principal: float
    interest_rate: float = 0.0
    start_date: str
    notes: Optional[str] = ""

class Loan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    person_name: str
    loan_type: str  # given, taken
    principal: float
    interest_rate: float = 0.0
    start_date: str
    notes: str = ""
    account_id: Optional[str] = None  # Linked account
    total_repaid: float = 0.0
    interest_paid: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class LoanUpdate(BaseModel):
    person_name: Optional[str] = None
    principal: Optional[float] = None
    interest_rate: Optional[float] = None
    start_date: Optional[str] = None
    notes: Optional[str] = None

class SettingsUpdate(BaseModel):
    current_password: str
    new_password: str

class TagPattern(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    pattern: str
    category_id: Optional[str] = None
    payee_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ================== AUTH ==================

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def generate_token() -> str:
    return hashlib.sha256(f"{datetime.now(timezone.utc).isoformat()}{uuid.uuid4()}".encode()).hexdigest()

async def get_current_user(token: str = None):
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    session = await db.sessions.find_one({"token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Session expired. Please login again.")
    return session

async def create_default_data():
    """Create default categories and accounts"""
    # Default Expense Categories
    expense_categories = [
        {"name": "Personal", "type": "expense", "icon": "user", "color": "#6366f1"},
        {"name": "Food & Dining", "type": "expense", "icon": "utensils", "color": "#f59e0b"},
        {"name": "Transport", "type": "expense", "icon": "car", "color": "#3b82f6"},
        {"name": "Utilities", "type": "expense", "icon": "zap", "color": "#10b981"},
        {"name": "Shopping", "type": "expense", "icon": "shopping-bag", "color": "#ec4899"},
        {"name": "Entertainment", "type": "expense", "icon": "film", "color": "#8b5cf6"},
        {"name": "Health", "type": "expense", "icon": "heart", "color": "#ef4444"},
        {"name": "Education", "type": "expense", "icon": "book", "color": "#14b8a6"},
        {"name": "Rent", "type": "expense", "icon": "home", "color": "#f97316"},
        {"name": "Interest Paid", "type": "expense", "icon": "percent", "color": "#dc2626"},
        {"name": "Other Expense", "type": "expense", "icon": "more-horizontal", "color": "#6b7280"},
    ]
    
    # Default Income Categories
    income_categories = [
        {"name": "Salary", "type": "income", "icon": "briefcase", "color": "#22c55e"},
        {"name": "Interest Received", "type": "income", "icon": "percent", "color": "#10b981"},
        {"name": "Investment Returns", "type": "income", "icon": "trending-up", "color": "#06b6d4"},
        {"name": "Other Income", "type": "income", "icon": "more-horizontal", "color": "#84cc16"},
    ]
    
    # Create categories
    for cat_data in expense_categories + income_categories:
        existing = await db.categories.find_one({"name": cat_data["name"], "type": cat_data["type"]}, {"_id": 0})
        if not existing:
            cat = Category(**cat_data)
            await db.categories.insert_one(cat.model_dump())
    
    # Create some common sub-categories
    personal_cat = await db.categories.find_one({"name": "Personal", "type": "expense"}, {"_id": 0})
    if personal_cat:
        sub_cats = ["Uber/Ola", "Subscription", "Grooming", "Misc"]
        for sub_name in sub_cats:
            existing = await db.categories.find_one({"name": sub_name, "parent_id": personal_cat["id"]}, {"_id": 0})
            if not existing:
                sub = Category(name=sub_name, parent_id=personal_cat["id"], type="expense")
                await db.categories.insert_one(sub.model_dump())
    
    food_cat = await db.categories.find_one({"name": "Food & Dining", "type": "expense"}, {"_id": 0})
    if food_cat:
        sub_cats = ["Restaurants", "Groceries", "Zomato/Swiggy"]
        for sub_name in sub_cats:
            existing = await db.categories.find_one({"name": sub_name, "parent_id": food_cat["id"]}, {"_id": 0})
            if not existing:
                sub = Category(name=sub_name, parent_id=food_cat["id"], type="expense")
                await db.categories.insert_one(sub.model_dump())
    
    utilities_cat = await db.categories.find_one({"name": "Utilities", "type": "expense"}, {"_id": 0})
    if utilities_cat:
        sub_cats = ["Electricity", "Internet", "Water", "Gas", "Mobile Recharge"]
        for sub_name in sub_cats:
            existing = await db.categories.find_one({"name": sub_name, "parent_id": utilities_cat["id"]}, {"_id": 0})
            if not existing:
                sub = Category(name=sub_name, parent_id=utilities_cat["id"], type="expense")
                await db.categories.insert_one(sub.model_dump())
    
    # Default Cash account
    cash_account = await db.accounts.find_one({"name": "Cash", "account_type": "cash"}, {"_id": 0})
    if not cash_account:
        cash = Account(name="Cash", account_type="cash", description="Cash in hand")
        await db.accounts.insert_one(cash.model_dump())

@api_router.post("/auth/setup", response_model=TokenResponse)
async def setup_password(data: UserCreate):
    existing = await db.users.find_one({}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Password already set")
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "password_hash": hash_password(data.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Create default categories and accounts
    await create_default_data()
    
    token = generate_token()
    await db.sessions.insert_one({"token": token, "created_at": datetime.now(timezone.utc).isoformat()})
    
    return TokenResponse(token=token, message="Setup complete")

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: LoginRequest):
    user = await db.users.find_one({}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=400, detail="Please setup password first")
    
    if user["password_hash"] != hash_password(data.password):
        raise HTTPException(status_code=401, detail="Invalid password")
    
    token = generate_token()
    await db.sessions.insert_one({"token": token, "created_at": datetime.now(timezone.utc).isoformat()})
    
    return TokenResponse(token=token, message="Login successful")

@api_router.get("/auth/check")
async def check_auth():
    user = await db.users.find_one({}, {"_id": 0})
    return {"setup_required": user is None}

@api_router.post("/auth/logout")
async def logout(token: str):
    await db.sessions.delete_one({"token": token})
    return {"message": "Logged out"}

@api_router.post("/auth/change-password")
async def change_password(data: SettingsUpdate, token: str):
    await get_current_user(token)
    user = await db.users.find_one({}, {"_id": 0})
    
    if user["password_hash"] != hash_password(data.current_password):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    
    await db.users.update_one({}, {"$set": {"password_hash": hash_password(data.new_password)}})
    return {"message": "Password changed successfully"}

@api_router.post("/auth/reset-all-data")
async def reset_all_data(token: str):
    """Delete all data except user credentials - fresh start"""
    await get_current_user(token)
    
    # Drop all collections except users and sessions
    await db.accounts.drop()
    await db.categories.drop()
    await db.transactions.drop()
    await db.loans.drop()
    await db.tag_patterns.drop()
    
    # Re-create default data
    await create_default_data()
    
    return {"message": "All data reset successfully. Default categories created."}

# ================== CATEGORIES ==================

@api_router.post("/categories", response_model=Category)
async def create_category(data: CategoryCreate, token: str):
    await get_current_user(token)
    category = Category(**data.model_dump())
    await db.categories.insert_one(category.model_dump())
    return category

@api_router.get("/categories")
async def get_categories(token: str, type: Optional[str] = None, include_children: bool = True):
    await get_current_user(token)
    query = {}
    if type:
        query["type"] = type
    
    categories = await db.categories.find(query, {"_id": 0}).to_list(1000)
    
    if include_children:
        # Build hierarchical structure
        parent_cats = [c for c in categories if c.get("parent_id") is None]
        for parent in parent_cats:
            parent["children"] = [c for c in categories if c.get("parent_id") == parent["id"]]
        return parent_cats
    
    return categories

@api_router.get("/categories/flat")
async def get_categories_flat(token: str, type: Optional[str] = None):
    """Get all categories in flat list"""
    await get_current_user(token)
    query = {}
    if type:
        query["type"] = type
    categories = await db.categories.find(query, {"_id": 0}).to_list(1000)
    return categories

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, data: CategoryCreate, token: str):
    await get_current_user(token)
    await db.categories.update_one({"id": category_id}, {"$set": data.model_dump()})
    category = await db.categories.find_one({"id": category_id}, {"_id": 0})
    return category

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, token: str):
    await get_current_user(token)
    # Also delete children
    await db.categories.delete_many({"parent_id": category_id})
    await db.categories.delete_one({"id": category_id})
    return {"message": "Category deleted"}

# ================== ACCOUNTS ==================

@api_router.post("/accounts", response_model=Account)
async def create_account(data: AccountCreate, token: str):
    await get_current_user(token)
    account = Account(**data.model_dump(), current_balance=data.opening_balance)
    await db.accounts.insert_one(account.model_dump())
    return account

@api_router.get("/accounts", response_model=List[Account])
async def get_accounts(token: str, account_type: Optional[str] = None):
    await get_current_user(token)
    query = {}
    if account_type:
        query["account_type"] = account_type
    accounts = await db.accounts.find(query, {"_id": 0}).to_list(1000)
    return accounts

@api_router.get("/accounts/{account_id}", response_model=Account)
async def get_account(account_id: str, token: str):
    await get_current_user(token)
    account = await db.accounts.find_one({"id": account_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    return account

@api_router.put("/accounts/{account_id}", response_model=Account)
async def update_account(account_id: str, data: AccountCreate, token: str):
    await get_current_user(token)
    update_data = data.model_dump()
    # Don't update current_balance directly
    if "opening_balance" in update_data:
        # Recalculate current_balance based on opening_balance change
        old_account = await db.accounts.find_one({"id": account_id}, {"_id": 0})
        if old_account:
            diff = update_data["opening_balance"] - old_account["opening_balance"]
            update_data["current_balance"] = old_account["current_balance"] + diff
    
    await db.accounts.update_one({"id": account_id}, {"$set": update_data})
    account = await db.accounts.find_one({"id": account_id}, {"_id": 0})
    return account

@api_router.delete("/accounts/{account_id}")
async def delete_account(account_id: str, token: str):
    await get_current_user(token)
    await db.accounts.delete_one({"id": account_id})
    return {"message": "Account deleted"}

# ================== TRANSACTIONS ==================

@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(data: TransactionCreate, token: str):
    await get_current_user(token)
    transaction = Transaction(**data.model_dump())
    await db.transactions.insert_one(transaction.model_dump())
    
    # Update account balance
    if data.transaction_type == "expense":
        await db.accounts.update_one({"id": data.account_id}, {"$inc": {"current_balance": -data.amount}})
    elif data.transaction_type == "income":
        await db.accounts.update_one({"id": data.account_id}, {"$inc": {"current_balance": data.amount}})
    elif data.transaction_type == "transfer" and data.payee_id:
        # Debit from source, credit to destination
        await db.accounts.update_one({"id": data.account_id}, {"$inc": {"current_balance": -data.amount}})
        await db.accounts.update_one({"id": data.payee_id}, {"$inc": {"current_balance": data.amount}})
    
    # Save pattern for auto-tagging
    if data.category_id or data.payee_id:
        pattern = re.sub(r'\d+', '', data.description)[:50]
        existing = await db.tag_patterns.find_one({"pattern": pattern}, {"_id": 0})
        if not existing:
            tag_pattern = TagPattern(pattern=pattern, category_id=data.category_id, payee_id=data.payee_id)
            await db.tag_patterns.insert_one(tag_pattern.model_dump())
    
    return transaction

@api_router.get("/transactions", response_model=List[Transaction])
async def get_transactions(
    token: str,
    account_id: Optional[str] = None,
    category_id: Optional[str] = None,
    transaction_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    untagged: Optional[bool] = None,
    limit: int = 500
):
    await get_current_user(token)
    query = {}
    
    if account_id:
        query["$or"] = [{"account_id": account_id}, {"payee_id": account_id}]
    if category_id:
        # Include children categories
        children = await db.categories.find({"parent_id": category_id}, {"_id": 0}).to_list(100)
        child_ids = [c["id"] for c in children]
        query["category_id"] = {"$in": [category_id] + child_ids}
    if transaction_type:
        query["transaction_type"] = transaction_type
    if untagged:
        query["category_id"] = None
        query["payee_id"] = None
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(limit)
    return transactions

@api_router.get("/transactions/{transaction_id}", response_model=Transaction)
async def get_transaction(transaction_id: str, token: str):
    await get_current_user(token)
    transaction = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return transaction

@api_router.put("/transactions/{transaction_id}", response_model=Transaction)
async def update_transaction(transaction_id: str, data: TransactionUpdate, token: str):
    await get_current_user(token)
    
    original = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not original:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    # Handle balance changes
    if "amount" in update_data or "transaction_type" in update_data or "account_id" in update_data:
        old_amount = original["amount"]
        old_type = original["transaction_type"]
        old_account = original["account_id"]
        
        new_amount = update_data.get("amount", old_amount)
        new_type = update_data.get("transaction_type", old_type)
        new_account = update_data.get("account_id", old_account)
        
        # Reverse old effect
        if old_type == "expense":
            await db.accounts.update_one({"id": old_account}, {"$inc": {"current_balance": old_amount}})
        elif old_type == "income":
            await db.accounts.update_one({"id": old_account}, {"$inc": {"current_balance": -old_amount}})
        
        # Apply new effect
        if new_type == "expense":
            await db.accounts.update_one({"id": new_account}, {"$inc": {"current_balance": -new_amount}})
        elif new_type == "income":
            await db.accounts.update_one({"id": new_account}, {"$inc": {"current_balance": new_amount}})
    
    await db.transactions.update_one({"id": transaction_id}, {"$set": update_data})
    transaction = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    
    # Save new pattern
    if update_data.get("category_id") or update_data.get("payee_id"):
        pattern = re.sub(r'\d+', '', original["description"])[:50]
        await db.tag_patterns.delete_one({"pattern": pattern})
        tag_pattern = TagPattern(
            pattern=pattern, 
            category_id=update_data.get("category_id", original.get("category_id")),
            payee_id=update_data.get("payee_id", original.get("payee_id"))
        )
        await db.tag_patterns.insert_one(tag_pattern.model_dump())
    
    return transaction

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str, token: str):
    await get_current_user(token)
    transaction = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if transaction:
        # Reverse balance effect
        if transaction["transaction_type"] == "expense":
            await db.accounts.update_one({"id": transaction["account_id"]}, {"$inc": {"current_balance": transaction["amount"]}})
        elif transaction["transaction_type"] == "income":
            await db.accounts.update_one({"id": transaction["account_id"]}, {"$inc": {"current_balance": -transaction["amount"]}})
        elif transaction["transaction_type"] == "transfer" and transaction.get("payee_id"):
            await db.accounts.update_one({"id": transaction["account_id"]}, {"$inc": {"current_balance": transaction["amount"]}})
            await db.accounts.update_one({"id": transaction["payee_id"]}, {"$inc": {"current_balance": -transaction["amount"]}})
    
    await db.transactions.delete_one({"id": transaction_id})
    return {"message": "Transaction deleted"}

@api_router.post("/transactions/bulk-tag")
async def bulk_tag_transactions(data: BulkTagRequest, token: str):
    await get_current_user(token)
    
    update_data = {}
    if data.category_id:
        update_data["category_id"] = data.category_id
    if data.payee_id:
        update_data["payee_id"] = data.payee_id
    
    if update_data:
        await db.transactions.update_many(
            {"id": {"$in": data.transaction_ids}},
            {"$set": update_data}
        )
        
        # Save patterns
        transactions = await db.transactions.find({"id": {"$in": data.transaction_ids}}, {"_id": 0}).to_list(100)
        for txn in transactions:
            pattern = re.sub(r'\d+', '', txn["description"])[:50]
            existing = await db.tag_patterns.find_one({"pattern": pattern}, {"_id": 0})
            if not existing:
                tag_pattern = TagPattern(pattern=pattern, category_id=data.category_id, payee_id=data.payee_id)
                await db.tag_patterns.insert_one(tag_pattern.model_dump())
    
    return {"message": f"Tagged {len(data.transaction_ids)} transactions"}

# ================== BANK STATEMENT UPLOAD ==================

async def apply_auto_tags(transactions: List[Dict]) -> List[Dict]:
    patterns = await db.tag_patterns.find({}, {"_id": 0}).to_list(1000)
    
    for txn in transactions:
        clean_desc = re.sub(r'\d+', '', txn["description"])[:50]
        for pattern in patterns:
            if pattern["pattern"] in clean_desc:
                if pattern.get("category_id"):
                    txn["category_id"] = pattern["category_id"]
                if pattern.get("payee_id"):
                    txn["payee_id"] = pattern["payee_id"]
                break
    
    return transactions

@api_router.post("/upload/bank-statement")
async def upload_bank_statement(file: UploadFile = File(...), account_id: str = None, token: str = None):
    await get_current_user(token)
    
    content = await file.read()
    
    try:
        # Read HDFC bank statement
        df = pd.read_excel(io.BytesIO(content), header=None)
        
        # Find the header row
        header_row = None
        for idx, row in df.iterrows():
            row_str = ' '.join(str(x) for x in row.values if pd.notna(x))
            if 'Date' in row_str and 'Narration' in row_str:
                header_row = idx
                break
        
        if header_row is None:
            raise HTTPException(status_code=400, detail="Could not find transaction headers in file")
        
        df = pd.read_excel(io.BytesIO(content), header=header_row)
        df = df[~df['Date'].astype(str).str.contains(r'\*+', regex=True, na=True)]
        df = df.dropna(subset=['Date'])
        
        transactions = []
        for _, row in df.iterrows():
            try:
                date_val = row.get('Date', '')
                if pd.isna(date_val) or str(date_val).strip() == '':
                    continue
                
                if isinstance(date_val, str):
                    date_parts = date_val.split('/')
                    if len(date_parts) == 3:
                        day, month, year = date_parts
                        if len(year) == 2:
                            year = '20' + year
                        date_str = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    else:
                        continue
                else:
                    date_str = pd.to_datetime(date_val).strftime('%Y-%m-%d')
                
                narration = str(row.get('Narration', ''))
                withdrawal = row.get('Withdrawal Amt.', 0)
                deposit = row.get('Deposit Amt.', 0)
                reference = str(row.get('Chq./Ref.No.', ''))
                
                withdrawal = float(withdrawal) if pd.notna(withdrawal) else 0
                deposit = float(deposit) if pd.notna(deposit) else 0
                
                if withdrawal > 0:
                    txn = {
                        "id": str(uuid.uuid4()),
                        "date": date_str,
                        "description": narration,
                        "amount": withdrawal,
                        "account_id": account_id or "",
                        "category_id": None,
                        "payee_id": None,
                        "transaction_type": "expense",
                        "reference": reference,
                        "source": "bank_import",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    transactions.append(txn)
                
                if deposit > 0:
                    txn = {
                        "id": str(uuid.uuid4()),
                        "date": date_str,
                        "description": narration,
                        "amount": deposit,
                        "account_id": account_id or "",
                        "category_id": None,
                        "payee_id": None,
                        "transaction_type": "income",
                        "reference": reference,
                        "source": "bank_import",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    transactions.append(txn)
                    
            except Exception as e:
                logging.error(f"Error parsing row: {e}")
                continue
        
        # Apply auto-tags
        transactions = await apply_auto_tags(transactions)
        
        return {"transactions": transactions, "count": len(transactions)}
        
    except Exception as e:
        logging.error(f"Error parsing file: {e}")
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")

@api_router.post("/upload/save-transactions")
async def save_uploaded_transactions(transactions: List[Dict[str, Any]], token: str):
    await get_current_user(token)
    
    if not transactions:
        return {"message": "No transactions to save", "count": 0}
    
    for txn in transactions:
        if "_id" in txn:
            del txn["_id"]
        await db.transactions.insert_one(txn)
        
        # Update account balance
        if txn.get("account_id"):
            if txn["transaction_type"] == "expense":
                await db.accounts.update_one({"id": txn["account_id"]}, {"$inc": {"current_balance": -txn["amount"]}})
            elif txn["transaction_type"] == "income":
                await db.accounts.update_one({"id": txn["account_id"]}, {"$inc": {"current_balance": txn["amount"]}})
        
        # Save tag pattern
        if txn.get("category_id") or txn.get("payee_id"):
            pattern = re.sub(r'\d+', '', txn["description"])[:50]
            existing = await db.tag_patterns.find_one({"pattern": pattern}, {"_id": 0})
            if not existing:
                tag_pattern = TagPattern(pattern=pattern, category_id=txn.get("category_id"), payee_id=txn.get("payee_id"))
                await db.tag_patterns.insert_one(tag_pattern.model_dump())
    
    return {"message": f"Saved {len(transactions)} transactions", "count": len(transactions)}

# ================== LOANS ==================

@api_router.post("/loans", response_model=Loan)
async def create_loan(data: LoanCreate, token: str):
    await get_current_user(token)
    
    # Create associated account
    account_type = "loan_receivable" if data.loan_type == "given" else "loan_payable"
    account = Account(
        name=f"Loan - {data.person_name}",
        account_type=account_type,
        description=f"Loan {data.loan_type} to/from {data.person_name}",
        person_name=data.person_name,
        opening_balance=data.principal,
        current_balance=data.principal
    )
    await db.accounts.insert_one(account.model_dump())
    
    loan_data = data.model_dump()
    loan_data['account_id'] = account.id
    loan = Loan(**loan_data)
    await db.loans.insert_one(loan.model_dump())
    return loan

@api_router.get("/loans", response_model=List[Loan])
async def get_loans(token: str, loan_type: Optional[str] = None):
    await get_current_user(token)
    query = {}
    if loan_type:
        query["loan_type"] = loan_type
    loans = await db.loans.find(query, {"_id": 0}).to_list(1000)
    return loans

@api_router.get("/loans/{loan_id}", response_model=Loan)
async def get_loan(loan_id: str, token: str):
    await get_current_user(token)
    loan = await db.loans.find_one({"id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    return loan

@api_router.get("/loans/{loan_id}/interest")
async def calculate_loan_interest(loan_id: str, token: str, as_of_date: Optional[str] = None):
    await get_current_user(token)
    loan = await db.loans.find_one({"id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    if loan["interest_rate"] == 0:
        return {"accrued_interest": 0, "total_due": loan["principal"] - loan["total_repaid"]}
    
    start = datetime.strptime(loan["start_date"], "%Y-%m-%d")
    end = datetime.strptime(as_of_date, "%Y-%m-%d") if as_of_date else datetime.now(timezone.utc).replace(tzinfo=None)
    days_elapsed = (end - start).days
    
    principal = loan["principal"]
    rate = loan["interest_rate"] / 100
    
    # Simple interest
    accrued_interest = principal * rate * (days_elapsed / 365)
    outstanding_principal = loan["principal"] - loan["total_repaid"]
    total_due = outstanding_principal + accrued_interest - loan["interest_paid"]
    
    return {
        "principal": loan["principal"],
        "outstanding_principal": outstanding_principal,
        "interest_rate": loan["interest_rate"],
        "days_elapsed": days_elapsed,
        "accrued_interest": round(accrued_interest, 2),
        "interest_paid": loan["interest_paid"],
        "interest_due": round(max(0, accrued_interest - loan["interest_paid"]), 2),
        "total_due": round(max(0, total_due), 2)
    }

@api_router.put("/loans/{loan_id}")
async def update_loan(loan_id: str, data: LoanUpdate, token: str):
    await get_current_user(token)
    loan = await db.loans.find_one({"id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.loans.update_one({"id": loan_id}, {"$set": update_data})
        if "person_name" in update_data and loan.get("account_id"):
            await db.accounts.update_one(
                {"id": loan["account_id"]},
                {"$set": {"name": f"Loan - {update_data['person_name']}", "person_name": update_data['person_name']}}
            )
    
    updated_loan = await db.loans.find_one({"id": loan_id}, {"_id": 0})
    return updated_loan

@api_router.delete("/loans/{loan_id}")
async def delete_loan(loan_id: str, token: str):
    await get_current_user(token)
    loan = await db.loans.find_one({"id": loan_id}, {"_id": 0})
    if loan and loan.get("account_id"):
        await db.accounts.delete_one({"id": loan["account_id"]})
    await db.loans.delete_one({"id": loan_id})
    return {"message": "Loan deleted"}

# ================== REPORTS ==================

@api_router.get("/reports/dashboard")
async def get_dashboard(token: str):
    await get_current_user(token)
    
    accounts = await db.accounts.find({}, {"_id": 0}).to_list(1000)
    
    bank_balance = sum(a["current_balance"] for a in accounts if a["account_type"] == "bank")
    cash_balance = sum(a["current_balance"] for a in accounts if a["account_type"] == "cash")
    loans_receivable = sum(a["current_balance"] for a in accounts if a["account_type"] == "loan_receivable")
    loans_payable = sum(a["current_balance"] for a in accounts if a["account_type"] == "loan_payable")
    investments = sum(a["current_balance"] for a in accounts if a["account_type"] == "investment")
    credit_cards = sum(a["current_balance"] for a in accounts if a["account_type"] == "credit_card")
    
    total_assets = bank_balance + cash_balance + loans_receivable + investments
    total_liabilities = loans_payable + credit_cards
    net_worth = total_assets - total_liabilities
    
    recent_transactions = await db.transactions.find({}, {"_id": 0}).sort("date", -1).to_list(10)
    
    # Enrich with category names
    categories = await db.categories.find({}, {"_id": 0}).to_list(1000)
    cat_map = {c["id"]: c["name"] for c in categories}
    for txn in recent_transactions:
        if txn.get("category_id"):
            txn["category_name"] = cat_map.get(txn["category_id"], "")
    
    # Monthly income/expense
    current_month = datetime.now(timezone.utc).strftime('%Y-%m')
    month_transactions = await db.transactions.find({"date": {"$regex": f"^{current_month}"}}, {"_id": 0}).to_list(10000)
    
    monthly_income = sum(t["amount"] for t in month_transactions if t["transaction_type"] == "income")
    monthly_expense = sum(t["amount"] for t in month_transactions if t["transaction_type"] == "expense")
    
    return {
        "bank_balance": bank_balance,
        "cash_balance": cash_balance,
        "loans_receivable": loans_receivable,
        "loans_payable": loans_payable,
        "investments": investments,
        "credit_cards": credit_cards,
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "net_worth": net_worth,
        "monthly_income": monthly_income,
        "monthly_expense": monthly_expense,
        "recent_transactions": recent_transactions
    }

@api_router.get("/reports/balance-sheet")
async def get_balance_sheet(token: str):
    await get_current_user(token)
    
    accounts = await db.accounts.find({}, {"_id": 0}).to_list(1000)
    
    assets = {
        "bank": [a for a in accounts if a["account_type"] == "bank"],
        "cash": [a for a in accounts if a["account_type"] == "cash"],
        "loans_receivable": [a for a in accounts if a["account_type"] == "loan_receivable"],
        "investments": [a for a in accounts if a["account_type"] == "investment"],
    }
    
    liabilities = {
        "loans_payable": [a for a in accounts if a["account_type"] == "loan_payable"],
        "credit_cards": [a for a in accounts if a["account_type"] == "credit_card"],
    }
    
    total_assets = sum(a["current_balance"] for a in accounts if a["account_type"] in ["bank", "cash", "loan_receivable", "investment"])
    total_liabilities = sum(a["current_balance"] for a in accounts if a["account_type"] in ["loan_payable", "credit_card"])
    
    return {
        "assets": assets,
        "liabilities": liabilities,
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "net_worth": total_assets - total_liabilities
    }

@api_router.get("/reports/income-expense")
async def get_income_expense(token: str, start_date: Optional[str] = None, end_date: Optional[str] = None):
    await get_current_user(token)
    
    query = {}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    categories = await db.categories.find({}, {"_id": 0}).to_list(1000)
    cat_map = {c["id"]: c for c in categories}
    
    # Group by category
    income_by_category = {}
    expense_by_category = {}
    
    for txn in transactions:
        cat_id = txn.get("category_id")
        cat = cat_map.get(cat_id, {"name": "Uncategorized", "parent_id": None})
        cat_name = cat.get("name", "Uncategorized")
        
        # If it's a sub-category, include parent
        if cat.get("parent_id"):
            parent = cat_map.get(cat["parent_id"])
            if parent:
                cat_name = f"{parent['name']} > {cat_name}"
        
        if txn["transaction_type"] == "income":
            income_by_category[cat_name] = income_by_category.get(cat_name, 0) + txn["amount"]
        elif txn["transaction_type"] == "expense":
            expense_by_category[cat_name] = expense_by_category.get(cat_name, 0) + txn["amount"]
    
    total_income = sum(income_by_category.values())
    total_expense = sum(expense_by_category.values())
    
    return {
        "income_by_category": income_by_category,
        "expense_by_category": expense_by_category,
        "total_income": total_income,
        "total_expense": total_expense,
        "net_income": total_income - total_expense
    }

@api_router.get("/reports/category/{category_id}")
async def get_category_report(category_id: str, token: str, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Get all transactions for a category and its children"""
    await get_current_user(token)
    
    # Get category and its children
    children = await db.categories.find({"parent_id": category_id}, {"_id": 0}).to_list(100)
    child_ids = [c["id"] for c in children]
    all_ids = [category_id] + child_ids
    
    query = {"category_id": {"$in": all_ids}}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    total = sum(t["amount"] for t in transactions)
    
    return {
        "transactions": transactions,
        "total": total,
        "count": len(transactions)
    }

# ================== EXPORT ==================

@api_router.get("/export/transactions")
async def export_transactions(token: str, start_date: Optional[str] = None, end_date: Optional[str] = None):
    await get_current_user(token)
    
    query = {}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    accounts = await db.accounts.find({}, {"_id": 0}).to_list(1000)
    categories = await db.categories.find({}, {"_id": 0}).to_list(1000)
    
    account_map = {a["id"]: a["name"] for a in accounts}
    cat_map = {c["id"]: c["name"] for c in categories}
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    
    headers = ["Date", "Description", "Amount", "Type", "Account", "Category", "Reference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for row, txn in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=txn["date"])
        ws.cell(row=row, column=2, value=txn["description"])
        ws.cell(row=row, column=3, value=txn["amount"])
        ws.cell(row=row, column=4, value=txn["transaction_type"])
        ws.cell(row=row, column=5, value=account_map.get(txn.get("account_id"), ""))
        ws.cell(row=row, column=6, value=cat_map.get(txn.get("category_id"), ""))
        ws.cell(row=row, column=7, value=txn.get("reference", ""))
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transactions.xlsx"}
    )

# ================== TAG PATTERNS ==================

@api_router.get("/tag-patterns")
async def get_tag_patterns(token: str):
    await get_current_user(token)
    patterns = await db.tag_patterns.find({}, {"_id": 0}).to_list(1000)
    return patterns

@api_router.delete("/tag-patterns/{pattern_id}")
async def delete_tag_pattern(pattern_id: str, token: str):
    await get_current_user(token)
    await db.tag_patterns.delete_one({"id": pattern_id})
    return {"message": "Pattern deleted"}

# ================== LEGACY ENDPOINTS (for backward compatibility) ==================

@api_router.get("/ledgers")
async def get_ledgers_legacy(token: str, type: Optional[str] = None, category: Optional[str] = None):
    """Legacy endpoint - redirects to accounts"""
    return await get_accounts(token, account_type=category)

@api_router.post("/ledgers")
async def create_ledger_legacy(data: Dict[str, Any], token: str):
    """Legacy endpoint - creates account"""
    await get_current_user(token)
    account_data = AccountCreate(
        name=data.get("name", ""),
        account_type=data.get("category", "bank"),
        opening_balance=data.get("opening_balance", 0),
        description=data.get("description", ""),
        person_name=data.get("person_name")
    )
    return await create_account(account_data, token)

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
